"""Customer list import: CSV/Excel parsing and phone column detection.

Requirement 2.1: each upload becomes a batch; column names are not fixed,
the whole original row is preserved, and the operator confirms the phone
column from the preview before scheduling may start.
"""

from __future__ import annotations

import csv
import io
import re
import secrets
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

MAX_ROWS = 10_000
PHONE_NAME_HINTS = ("电话", "手机", "号码", "phone", "mobile", "tel")

_MOBILE_PATTERN = re.compile(r"^1[3-9]\d{9}$")
_LANDLINE_PATTERN = re.compile(r"^0\d{9,11}$")


class ImportParseError(ValueError):
    """Raised when an uploaded list cannot be turned into a batch."""


@dataclass(frozen=True)
class ParsedRow:
    row_number: int  # spreadsheet row, header occupies row 1
    data: dict[str, str] = field(default_factory=dict)


@dataclass(frozen=True)
class ParsedTable:
    columns: list[str]
    rows: list[ParsedRow]

    @property
    def total(self) -> int:
        return len(self.rows)


def new_batch_id(now: datetime | None = None) -> str:
    moment = now or datetime.now(timezone.utc)
    return f"B-{moment.strftime('%Y%m%d')}-{secrets.randbelow(10000):04d}"


_GENDER_NAME_HINTS = re.compile(r"性别|性別|gender|sex", re.IGNORECASE)


def extract_gender(raw_data: dict[str, object] | None) -> str:
    """Normalize the imported gender field to ``"male"``/``"female"``/``""``."""
    if not isinstance(raw_data, dict):
        return ""
    for key, value in raw_data.items():
        if not _GENDER_NAME_HINTS.search(str(key)):
            continue
        text = str(value or "").strip()
        if "男" in text or text.lower() in {"male", "m"}:
            return "male"
        if "女" in text or text.lower() in {"female", "f"}:
            return "female"
    return ""


def parse_table(filename: str, content: bytes) -> ParsedTable:
    """Dispatch on the file extension and enforce the row bound."""
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        table = _parse_csv(content)
    elif suffix == ".xlsx":
        table = _parse_excel(content)
    elif suffix == ".xls":
        raise ImportParseError("不支持旧版 .xls 格式，请另存为 .xlsx 或 CSV 后重试")
    else:
        raise ImportParseError("仅支持 .csv 或 .xlsx 文件")
    if table.total == 0:
        raise ImportParseError("清单中没有有效数据行")
    if table.total > MAX_ROWS:
        raise ImportParseError(f"单批最多 {MAX_ROWS} 行，当前 {table.total} 行")
    return table


def _clean_columns(header: list[str]) -> list[str]:
    """Trim header cells and drop trailing empty columns."""
    columns = [str(cell).strip() for cell in header]
    while columns and not columns[-1]:
        columns.pop()
    if not columns:
        raise ImportParseError("清单缺少列名（首行为空）")
    return columns


def _row_data(columns: list[str], cells: list[str]) -> dict[str, str]:
    data: dict[str, str] = {}
    for index, column in enumerate(columns):
        data[column] = str(cells[index]).strip() if index < len(cells) else ""
    return data


def _is_blank(cells: list[str]) -> bool:
    return all(not str(cell).strip() for cell in cells)


def _parse_csv(content: bytes) -> ParsedTable:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise ImportParseError("CSV 不是 UTF-8 编码，请用 UTF-8 另存后重试") from error
    reader = csv.reader(io.StringIO(text))
    rows_iter = list(reader)
    if not rows_iter:
        raise ImportParseError("清单中没有有效数据行")
    columns = _clean_columns(rows_iter[0])
    rows: list[ParsedRow] = []
    for row_number, cells in enumerate(rows_iter[1:], start=2):
        if _is_blank(cells):
            continue
        rows.append(ParsedRow(row_number=row_number, data=_row_data(columns, cells)))
    return ParsedTable(columns=columns, rows=rows)


def _cell_text(value: object) -> str:
    """Excel cells may be numeric; render integral floats without '.0' so
    phone numbers survive as plain digit strings."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_excel(content: bytes) -> ParsedTable:
    try:
        from openpyxl import load_workbook
    except ImportError as error:  # pragma: no cover - dependency is required
        raise ImportParseError("服务端缺少 openpyxl，无法解析 Excel") from error
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as error:
        raise ImportParseError("Excel 文件损坏或格式不正确") from error
    try:
        sheet = workbook.worksheets[0]
        all_rows = [
            [_cell_text(cell.value) for cell in row] for row in sheet.iter_rows()
        ]
    finally:
        workbook.close()
    if not all_rows:
        raise ImportParseError("清单中没有有效数据行")
    columns = _clean_columns(all_rows[0])
    rows: list[ParsedRow] = []
    for row_number, cells in enumerate(all_rows[1:], start=2):
        if _is_blank(cells):
            continue
        rows.append(ParsedRow(row_number=row_number, data=_row_data(columns, cells)))
    return ParsedTable(columns=columns, rows=rows)


def normalize_phone(value: str) -> str:
    return re.sub(r"[\s\-]", "", value)


def looks_like_phone(value: str) -> bool:
    normalized = normalize_phone(value)
    return bool(_MOBILE_PATTERN.match(normalized) or _LANDLINE_PATTERN.match(normalized))


def detect_phone_column(table: ParsedTable) -> str | None:
    """Priority 1: column name hints. Priority 2: a column whose non-empty
    values are all phone-shaped. Returns None when nothing matches."""
    for column in table.columns:
        lowered = column.lower()
        if any(hint in lowered for hint in PHONE_NAME_HINTS):
            return column
    for column in table.columns:
        values = [row.data.get(column, "") for row in table.rows]
        values = [value for value in values if value]
        if values and all(looks_like_phone(value) for value in values):
            return column
    return None
